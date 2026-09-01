import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.models import Ticket
from app.services.insights_service import obtener_insight
from app.services.reporte_service import calcular_metricas, rango_periodo, validar_periodo

MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]

ENCABEZADO_AZUL = PatternFill("solid", fgColor="1D44AA")


def _nombre_periodo(periodo: str) -> str:
    return f"{MESES_ES[int(periodo[5:7]) - 1]} {periodo[:4]}"


def _tickets_del_periodo(db: Session, periodo: str) -> list[Ticket]:
    inicio, fin = rango_periodo(periodo)
    return list(
        db.scalars(
            select(Ticket)
            .options(joinedload(Ticket.cliente))
            .where(Ticket.fecha_creacion >= inicio, Ticket.fecha_creacion < fin)
            .order_by(Ticket.fecha_creacion)
        ).all()
    )


def _seccion_hoja(hoja, fila: int, titulo: str, pares: list[tuple[str, object]]) -> int:
    hoja.cell(row=fila, column=1, value=titulo).font = Font(bold=True, color="1D44AA")
    fila += 1
    for clave, valor in pares:
        hoja.cell(row=fila, column=1, value=clave)
        celda = hoja.cell(row=fila, column=2, value=valor)
        if isinstance(valor, (int, float)):
            celda.alignment = Alignment(horizontal="right")
        fila += 1
    return fila + 1


def generar_excel(db: Session, periodo: str) -> tuple[bytes, str]:
    validar_periodo(periodo)
    metricas = calcular_metricas(db, periodo)

    wb = Workbook()
    hoja = wb.active
    hoja.title = "Resumen"
    hoja.column_dimensions["A"].width = 30
    hoja.column_dimensions["B"].width = 14

    hoja["A1"] = f"Reporte mensual de atencion al cliente - {_nombre_periodo(periodo)}"
    hoja["A1"].font = Font(bold=True, size=13)

    fila = _seccion_hoja(
        hoja,
        3,
        "Indicadores generales",
        [
            ("Total de tickets", metricas["total_tickets"]),
            ("Resueltos por IA", metricas["resueltos_por_ia"]),
            ("Resueltos por agente", metricas["resueltos_manual"]),
            ("Escalados a revision manual", metricas["escalados"]),
            ("Pendientes", metricas["pendientes"]),
            ("Tasa de resolucion IA (%)", metricas["tasa_resolucion_ia_pct"]),
            (
                "Tiempo promedio de atencion (s)",
                metricas["tiempo_promedio_atencion_seg"] if metricas["tiempo_promedio_atencion_seg"] is not None else "",
            ),
            ("Satisfaccion promedio", metricas["satisfaccion_promedio"] if metricas["satisfaccion_promedio"] is not None else ""),
        ],
    )

    ETIQUETAS_TIPO = {"consulta": "Consultas", "solicitud": "Solicitudes", "queja": "Quejas"}
    ETIQUETAS_ESTADO = {
        "abierto": "Abiertos",
        "escalado": "Escalados",
        "resuelto_ia": "Resueltos por IA",
        "resuelto_manual": "Resueltos por agente",
        "cerrado": "Cerrados",
    }
    ETIQUETAS_PRIORIDAD = {"baja": "Baja", "media": "Media", "alta": "Alta", "critica": "Critica"}

    fila = _seccion_hoja(
        hoja,
        fila,
        "Distribucion por tipo",
        [(ETIQUETAS_TIPO[k], v) for k, v in metricas["por_tipo"].items()],
    )
    fila = _seccion_hoja(
        hoja,
        fila,
        "Distribucion por estado",
        [(ETIQUETAS_ESTADO[k], v) for k, v in metricas["por_estado"].items()],
    )
    fila = _seccion_hoja(
        hoja,
        fila,
        "Distribucion por prioridad",
        [(ETIQUETAS_PRIORIDAD[k], v) for k, v in metricas["por_prioridad"].items()],
    )

    hoja_tickets = wb.create_sheet("Tickets")
    encabezados = [
        "Codigo", "Cliente", "Email", "Tipo", "Estado", "Prioridad",
        "Resuelto por IA", "Tiempo (s)", "Fecha creacion", "Fecha resolucion",
    ]
    ancho_columnas = [16, 24, 28, 12, 14, 11, 15, 11, 18, 18]
    for i, encabezado in enumerate(encabezados, start=1):
        celda = hoja_tickets.cell(row=1, column=i, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = ENCABEZADO_AZUL
        hoja_tickets.column_dimensions[chr(64 + i)].width = ancho_columnas[i - 1]

    tickets = _tickets_del_periodo(db, periodo)
    for n, t in enumerate(tickets, start=2):
        valores = [
            t.codigo,
            t.cliente.nombre,
            t.cliente.email,
            t.tipo.value,
            t.estado.value,
            t.prioridad.value,
            "Si" if t.resuelto_por_ia else "No",
            t.tiempo_atencion_seg,
            t.fecha_creacion.strftime("%Y-%m-%d %H:%M"),
            t.fecha_resolucion.strftime("%Y-%m-%d %H:%M") if t.fecha_resolucion else "",
        ]
        for i, valor in enumerate(valores, start=1):
            hoja_tickets.cell(row=n, column=i, value=valor)

    hoja_insights = wb.create_sheet("Insights IA")
    hoja_insights.column_dimensions["A"].width = 110
    insight = obtener_insight(db, periodo)
    if insight is None:
        hoja_insights["A1"] = (
            "Aun no se ha generado el resumen ejecutivo con IA para este periodo."
        )
    else:
        hoja_insights["A1"] = f"Resumen ejecutivo ({insight.modelo})"
        hoja_insights["A1"].font = Font(bold=True, color="1D44AA")
        hoja_insights["A2"] = insight.resumen
        fila = 4
        hoja_insights.cell(row=fila, column=1, value="Hallazgos").font = Font(bold=True)
        for h in insight.hallazgos:
            fila += 1
            detalle = f" - {h['detalle']}" if h.get("detalle") else ""
            hoja_insights.cell(row=fila, column=1, value=f"{h['titulo']}{detalle}")
        fila += 2
        hoja_insights.cell(row=fila, column=1, value="Recomendaciones").font = Font(bold=True)
        for r in insight.recomendaciones:
            fila += 1
            detalle = f" - {r['detalle']}" if r.get("detalle") else ""
            hoja_insights.cell(row=fila, column=1, value=f"{r['titulo']}{detalle}")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), f"reporte-{periodo}.xlsx"


def generar_pdf(db: Session, periodo: str) -> tuple[bytes, str]:
    validar_periodo(periodo)
    metricas = calcular_metricas(db, periodo)

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        topMargin=1.6 * cm,
        bottomMargin=1.6 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        title=f"Reporte mensual {_nombre_periodo(periodo)}",
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("TituloEs", parent=estilos["Title"], fontSize=17, textColor=colors.HexColor("#1D44AA"))
    estilo_seccion = ParagraphStyle("SeccionEs", parent=estilos["Heading2"], fontSize=13, spaceBefore=14, textColor=colors.HexColor("#1D44AA"))
    estilo_normal = estilos["BodyText"]
    estilo_nota = ParagraphStyle("NotaEs", parent=estilos["Italic"], fontSize=9, textColor=colors.HexColor("#64748B"))

    elementos = [
        Paragraph(f"Reporte mensual de atencion al cliente", estilo_titulo),
        Paragraph(_nombre_periodo(periodo).capitalize(), ParagraphStyle("Sub", parent=estilos["Normal"], alignment=1, spaceAfter=10)),
        Paragraph("Indicadores generales", estilo_seccion),
    ]

    filas = [["Indicador", "Valor"]]
    for clave, valor in [
        ("Total de tickets", metricas["total_tickets"]),
        ("Resueltos por IA", metricas["resueltos_por_ia"]),
        ("Resueltos por agente", metricas["resueltos_manual"]),
        ("Escalados a revision manual", metricas["escalados"]),
        ("Pendientes", metricas["pendientes"]),
        ("Tasa de resolucion IA (%)", metricas["tasa_resolucion_ia_pct"]),
        ("Tiempo promedio de atencion (s)", metricas["tiempo_promedio_atencion_seg"] if metricas["tiempo_promedio_atencion_seg"] is not None else "-"),
        ("Satisfaccion promedio", metricas["satisfaccion_promedio"] if metricas["satisfaccion_promedio"] is not None else "-"),
    ]:
        filas.append([str(clave), str(valor)])

    tabla = Table(filas, colWidths=[10 * cm, 3.5 * cm])
    tabla.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D44AA")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ])
    )
    elementos.append(tabla)

    distribucion = [["Tipo / Estado / Prioridad", "Cantidad"]]
    for k, v in metricas["por_tipo"].items():
        distribucion.append([f"Tipo: {k}", str(v)])
    for k, v in metricas["por_estado"].items():
        distribucion.append([f"Estado: {k}", str(v)])
    for k, v in metricas["por_prioridad"].items():
        distribucion.append([f"Prioridad: {k}", str(v)])
    tabla_dist = Table(distribucion, colWidths=[10 * cm, 3.5 * cm])
    tabla_dist.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elementos += [Paragraph("Distribucion de solicitudes", estilo_seccion), tabla_dist]

    insight = obtener_insight(db, periodo)
    elementos.append(Paragraph("Resumen ejecutivo generado con IA", estilo_seccion))
    if insight is None:
        elementos.append(Paragraph("Aun no se ha generado el analisis con IA para este periodo.", estilo_nota))
    else:
        elementos.append(Paragraph(insight.resumen, estilo_normal))
        elementos.append(Spacer(1, 6))
        elementos.append(Paragraph("<b>Hallazgos</b>", estilo_normal))
        for h in insight.hallazgos:
            detalle = f" - {h['detalle']}" if h.get("detalle") else ""
            elementos.append(Paragraph(f"&bull; <b>{h['titulo']}.</b> {detalle}" if detalle else f"&bull; {h['titulo']}.", estilo_normal))
        elementos.append(Spacer(1, 6))
        elementos.append(Paragraph("<b>Recomendaciones</b>", estilo_normal))
        for r in insight.recomendaciones:
            detalle = f" - {r['detalle']}" if r.get("detalle") else ""
            elementos.append(Paragraph(f"&bull; <b>{r['titulo']}.</b> {detalle}" if detalle else f"&bull; {r['titulo']}.", estilo_normal))

    total_tickets = len(_tickets_del_periodo(db, periodo))
    elementos += [
        Spacer(1, 14),
        Paragraph(f"Documento generado automaticamente sobre {total_tickets} ticket(s) del periodo.", estilo_nota),
    ]

    doc.build(elementos)
    return bio.getvalue(), f"reporte-{periodo}.pdf"
